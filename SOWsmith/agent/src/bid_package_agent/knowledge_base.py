"""Load the SharePoint knowledge base from the local folders and provide retrieval.

Mirrors the production wiring:
  * Approved-Exemplars  -> generation grounding (Flow B / Topic 1)
  * Reference-Library   -> Q&A grounding        (Flow A / Topic 2)
  * Lists/Reviewers.xlsx -> reviewer routing    (Power Automate action 4)
"""
from __future__ import annotations

from dataclasses import dataclass
from functools import cached_property
from pathlib import Path

from .config import Config
from .footer import parse_footer
from .retrieval import TfidfIndex
from .textextract import extract_text

_GROUNDING_EXTS = {".docx", ".pdf", ".pptx", ".xlsx"}


@dataclass
class Document:
    path: Path
    title: str
    text: str
    library: str
    service_type: str | None = None
    review_status: str | None = None

    @property
    def filename(self) -> str:
        return self.path.name


@dataclass
class Reviewer:
    service_type: str
    reviewer_email: str
    legal_reviewer_email: str


def _pretty_stem(p: Path) -> str:
    import re as _re
    stem = _re.sub(r"_v\d.*$", "", p.stem)          # drop version/date suffix
    return _re.sub(r"[-_]+", " ", stem).strip()


def _derive_title(text: str, p: Path) -> str:
    # Spreadsheets/decks: the first extracted line is a sheet marker or kicker, so a
    # prettified filename reads better as a citation. Documents: use the first real line.
    if p.suffix.lower() in (".xlsx", ".pptx"):
        return _pretty_stem(p)
    for ln in text.splitlines():
        s = ln.strip()
        if s and not s.startswith("#") and len(s) > 4:
            return s
    return _pretty_stem(p)


def _load_library(folder: Path, library: str) -> list[Document]:
    docs: list[Document] = []
    if not folder.is_dir():
        return docs
    for p in sorted(folder.iterdir()):
        if p.suffix.lower() not in _GROUNDING_EXTS or p.name.startswith("_"):
            continue
        text = extract_text(p)
        title = _derive_title(text, p)
        ftr = parse_footer(text)
        docs.append(Document(
            path=p, title=title, text=text, library=library,
            service_type=ftr.service_type if ftr else None,
            review_status=ftr.review_status if ftr else None,
        ))
    return docs


def load_reviewers(xlsx: Path) -> dict[str, Reviewer]:
    out: dict[str, Reviewer] = {}
    if not xlsx.is_file():
        return out
    from openpyxl import load_workbook
    ws = load_workbook(str(xlsx), read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        pt = str(row[0]).strip()
        out[pt] = Reviewer(
            service_type=pt,
            reviewer_email=str(row[1] or "").strip(),
            legal_reviewer_email=str(row[2] or "").strip(),
        )
    return out


class KnowledgeBase:
    def __init__(self, cfg: Config):
        self.cfg = cfg
        self.exemplars = _load_library(cfg.exemplars_dir, "Approved-Exemplars")
        self.reference = _load_library(cfg.reference_dir, "Reference-Library")
        self.reviewers = load_reviewers(cfg.reviewers_xlsx)
        # Last-seen review status per service type, taken from the curated KB footers. This is
        # only a hint: a bid package's review tier is CONTENT-driven (the same service type can be
        # STANDARD or LEGAL depending on the commercial/legal terms), so generate_draft() decides
        # the tier from the request and uses this map only as a fallback for known service types.
        self._status_by_type: dict[str, str] = {}
        for d in self.exemplars + self.reference:
            if d.service_type and d.review_status:
                self._status_by_type.setdefault(d.service_type, d.review_status)

    @cached_property
    def _exemplar_index(self) -> TfidfIndex:
        idx = TfidfIndex()
        for d in self.exemplars:
            idx.add(d.filename, f"{d.service_type or ''} {d.title}\n{d.text}", payload=d)
        return idx.build()

    @cached_property
    def _reference_index(self) -> TfidfIndex:
        idx = TfidfIndex()
        for d in self.reference:
            idx.add(d.filename, f"{d.title}\n{d.text}", payload=d)
        return idx.build()

    def exemplars_for(self, service_type: str, intent: str, k: int = 2) -> list[Document]:
        # Prefer an exact service_type match, then fall back to semantic similarity.
        exact = [d for d in self.exemplars if (d.service_type or "").lower() == service_type.lower()]
        hits = [d.payload for _, d in self._exemplar_index.search(f"{service_type} {intent}", top_k=k + 2)]
        ordered = exact + [d for d in hits if d not in exact]
        return ordered[:k] if ordered else self.exemplars[:k]

    def search_reference(self, question: str, k: int = 3) -> list[tuple[float, Document]]:
        return [(s, d.payload) for s, d in self._reference_index.search(question, top_k=k)]

    def reviewer_for(self, service_type: str) -> Reviewer | None:
        return self.reviewers.get(service_type)

    def review_status_for(self, service_type: str) -> str | None:
        """The last-seen review status for a known service type (from KB footers), or None."""
        return self._status_by_type.get(service_type)
